"""Formatted Excel exporter for PA-FR-010.3."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_presentation_builder import format_cell_value

SHEET_NAMES = {
    "metadata": "Overview",
    "executive_summary": "Summary",
    "coverage_summary": "Coverage",
    "embedding_summary": "Embeddings",
    "cluster_summary": "Clusters",
    "redundancy_summary": "Redundancy",
    "similarity_summary": "Similarity",
    "correlation_summary": "Correlation",
    "validation": "Validation",
    "appendix": "Appendix",
}

TITLE_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FILL = PatternFill("solid", fgColor="E9EEF5")
STATUS_FILL = PatternFill("solid", fgColor="F4F6F9")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD2DC"),
    right=Side(style="thin", color="CBD2DC"),
    top=Side(style="thin", color="CBD2DC"),
    bottom=Side(style="thin", color="CBD2DC"),
)


def _excel_value(value: Any) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", format_cell_value(value))
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _append_row(sheet: Any, values: Sequence[Any], *, style: str = "body") -> int:
    sheet.append([_excel_value(value) for value in values])
    row_number = sheet.max_row
    for cell in sheet[row_number]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if style == "title":
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=14)
        elif style == "header":
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        elif style == "status":
            cell.fill = STATUS_FILL
            cell.font = Font(bold=True)
        else:
            cell.border = THIN_BORDER
    return row_number


def _append_table(sheet: Any, table: Mapping[str, Any]) -> None:
    sheet.append([])
    _append_row(sheet, [table.get("title") or "Summary"], style="status")
    columns = list(table.get("columns") or [])
    rows = list(table.get("rows") or [])
    if not columns:
        _append_row(sheet, ["No columns available."])
        return
    _append_row(
        sheet,
        [str(column).replace("_", " ").title() for column in columns],
        style="header",
    )
    if not rows:
        _append_row(sheet, ["No rows available."])
        return
    for row in rows:
        _append_row(sheet, [row.get(column) for column in columns])


def _chart_rows(chart: Mapping[str, Any]) -> tuple[list[str], list[Mapping[str, Any]]]:
    data = chart.get("data")
    if isinstance(data, Mapping):
        return (
            ["label", "value"],
            [
                {"label": key, "value": value}
                for key, value in sorted(data.items(), key=lambda item: str(item[0]))
            ],
        )
    if isinstance(data, Sequence) and not isinstance(data, (str, bytes)):
        rows = [item for item in data if isinstance(item, Mapping)]
        columns: list[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(str(key))
        return columns, rows
    return [], []


def _append_section(sheet: Any, section: Mapping[str, Any]) -> None:
    _append_row(sheet, [section.get("title") or "Section"], style="title")
    _append_row(
        sheet,
        ["Section Status", section.get("status") or "Missing"],
        style="status",
    )
    kpis = list(section.get("kpis") or [])
    if kpis:
        sheet.append([])
        _append_row(sheet, ["KPI", "Value"], style="header")
        for item in kpis:
            _append_row(sheet, [item.get("label"), item.get("display")])
    messages = list(section.get("messages") or [])
    if messages:
        sheet.append([])
        _append_row(sheet, ["Warnings"], style="header")
        for message in messages:
            _append_row(sheet, [message])
    for chart in section.get("charts") or []:
        columns, rows = _chart_rows(chart)
        if columns:
            _append_table(
                sheet,
                {
                    "title": f"Chart Data: {chart.get('title') or 'Chart'}",
                    "columns": columns,
                    "rows": rows,
                },
            )
    for table in section.get("tables") or []:
        _append_table(sheet, table)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _autosize(sheet: Any) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value or "")
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), longest_line)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 12),
            60,
        )


def _model_datetime(presentation: Mapping[str, Any]) -> datetime:
    value = str((presentation.get("provenance") or {}).get("generation_timestamp") or "")
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None)
    except ValueError:
        return datetime(2000, 1, 1)


def export_excel(presentation: Mapping[str, Any]) -> bytes:
    """Return a readable workbook containing every report row."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.iso_dates = True
    workbook.properties.title = "Pattern Quality Report"
    workbook.properties.subject = "PA-FR-010 Single Log Pattern Quality Report"
    workbook.properties.creator = "Pattern Analysis Agent"
    workbook.properties.lastModifiedBy = "Pattern Analysis Agent"
    model_datetime = _model_datetime(presentation)
    workbook.properties.created = model_datetime
    workbook.properties.modified = model_datetime

    sections = list(presentation.get("sections") or [])
    section_by_id = {
        str(section.get("id")): section
        for section in sections
    }
    ordered_ids = [str(item) for item in presentation.get("section_order") or []]

    for section_id in ordered_ids[:2]:
        section = section_by_id.get(section_id)
        if section is None:
            continue
        sheet = workbook.create_sheet(SHEET_NAMES.get(section_id, section_id[:31]))
        _append_section(sheet, section)
        _autosize(sheet)

    kpi_sheet = workbook.create_sheet("KPIs")
    _append_row(kpi_sheet, ["Report KPIs"], style="title")
    _append_row(kpi_sheet, ["Section", "KPI", "Value"], style="header")
    for section_id in ordered_ids:
        section = section_by_id.get(section_id) or {}
        for item in section.get("kpis") or []:
            _append_row(
                kpi_sheet,
                [section.get("title"), item.get("label"), item.get("display")],
            )
    kpi_sheet.freeze_panes = "A3"
    kpi_sheet.sheet_view.showGridLines = False
    _autosize(kpi_sheet)

    for section_id in ordered_ids[2:]:
        section = section_by_id.get(section_id)
        if section is None:
            continue
        sheet = workbook.create_sheet(SHEET_NAMES.get(section_id, section_id[:31]))
        _append_section(sheet, section)
        _autosize(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
