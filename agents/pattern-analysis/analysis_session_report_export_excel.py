"""Deterministic Excel exporter for PA-FR-010.AS.3."""
from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping, Sequence
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_NAMES = {
    "overview": "Overview",
    "session_summary": "Session Summary",
    "coverage": "Coverage",
    "requirement_1_ingestion": "Req 1 Ingest Pattern Files",
    "requirement_2_vectors": "Req 2 Parse Patterns",
    "requirement_3_metadata": "Req 3 Extract Metadata",
    "requirement_4_toggle": "Req 4 Pattern Toggle",
    "embeddings": "Embeddings",
    "clustering": "Clustering",
    "redundancy": "Redundancy",
    "similarity": "Similarity",
    "pattern_outcomes": "Pattern Outcomes",
    "failure_risk_by_lot": "Failure Risk LOT",
    "failure_risk": "Failure Risk Log",
    "anomaly_by_lot": "Anomaly LOT",
    "anomaly": "Anomaly Log",
    "root_cause_by_lot": "Root Cause LOT",
    "root_cause": "Root Cause Log",
    "recommendations_by_lot": "Recommendations LOT",
    "recommendations": "Recommendations Log",
    "validation": "Validation",
    "appendix": "Appendix",
}

TITLE_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FILL = PatternFill("solid", fgColor="E9EEF5")
STATUS_FILL = PatternFill("solid", fgColor="F4F6F9")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD2DC"),
    right=Side(style="thin", color="CBD2DC"),
    top=Side(style="thin", color="CBD2DC"),
    bottom=Side(style="thin", color="CBD2DC"),
)
FIXED_ZIP_TIMESTAMP = (2000, 1, 1, 0, 0, 0)


def _format(value: Any) -> str:
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, tuple)):
        return ", ".join(_format(item) for item in value) or "—"
    if isinstance(value, Mapping):
        return json.dumps(value, sort_keys=True, ensure_ascii=False)
    return str(value)


def _excel_value(value: Any) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", _format(value))
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _append_row(sheet: Any, values: Sequence[Any], *, style: str = "body") -> None:
    sheet.append([_excel_value(value) for value in values])
    for cell in sheet[sheet.max_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if style == "title":
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=14)
        elif style == "header":
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        elif style == "status":
            cell.fill = STATUS_FILL
            cell.font = Font(bold=True)
        else:
            cell.border = THIN_BORDER


def _append_table(sheet: Any, table: Mapping[str, Any]) -> None:
    sheet.append([])
    _append_row(sheet, [table.get("title") or "Summary"], style="status")
    columns = list(table.get("columns") or [])
    rows = list(table.get("rows") or [])
    if not columns:
        _append_row(sheet, ["No columns available."])
        return
    _append_row(
        sheet,
        [str(column).replace("_", " ").title() for column in columns],
        style="header",
    )
    if not rows:
        _append_row(sheet, ["No rows available."])
        return
    for row in rows:
        _append_row(sheet, [row.get(column) for column in columns])


def _append_section(sheet: Any, section: Mapping[str, Any]) -> None:
    _append_row(sheet, [section.get("title") or "Section"], style="title")
    _append_row(
        sheet,
        ["Section Status", section.get("status") or "Missing"],
        style="status",
    )
    kpis = list(section.get("kpis") or [])
    if kpis:
        sheet.append([])
        _append_row(sheet, ["KPI", "Value"], style="header")
        for item in kpis:
            _append_row(sheet, [item.get("label"), item.get("display")])
    messages = list(section.get("messages") or [])
    if messages:
        sheet.append([])
        _append_row(sheet, ["Warnings"], style="header")
        for message in messages:
            _append_row(sheet, [message])
    for table in section.get("tables") or []:
        _append_table(sheet, table)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _autosize(sheet: Any) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value or "")
            longest = max((len(line) for line in value.splitlines()), default=0)
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 12),
            60,
        )


def _model_datetime(projection: Mapping[str, Any]) -> datetime:
    value = str(
        (projection.get("provenance") or {}).get("generation_timestamp") or ""
    )
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(
            tzinfo=None
        )
    except ValueError:
        return datetime(2000, 1, 1)


def _normalize_archive(content: bytes) -> bytes:
    source = BytesIO(content)
    target = BytesIO()
    with ZipFile(source, "r") as input_zip:
        with ZipFile(
            target,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
        ) as output_zip:
            for source_info in input_zip.infolist():
                data = input_zip.read(source_info.filename)
                if source_info.filename == "docProps/core.xml":
                    created = re.search(
                        rb"<dcterms:created[^>]*>(.*?)</dcterms:created>",
                        data,
                    )
                    if created:
                        data = re.sub(
                            rb"(<dcterms:modified[^>]*>).*?(</dcterms:modified>)",
                            rb"\g<1>" + created.group(1) + rb"\g<2>",
                            data,
                            count=1,
                        )
                target_info = ZipInfo(source_info.filename, FIXED_ZIP_TIMESTAMP)
                target_info.compress_type = ZIP_DEFLATED
                target_info.create_system = 0
                target_info.external_attr = source_info.external_attr
                target_info.internal_attr = source_info.internal_attr
                target_info.flag_bits = source_info.flag_bits
                output_zip.writestr(target_info, data)
    return target.getvalue()


def export_analysis_session_excel(projection: Mapping[str, Any]) -> bytes:
    """Return a deterministic workbook with one required sheet per section."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.iso_dates = True
    workbook.properties.title = "Analysis Session Quality Report"
    workbook.properties.subject = "PA-FR-010.AS Analysis Session Quality Report"
    workbook.properties.creator = "Pattern Analysis Agent"
    workbook.properties.lastModifiedBy = "Pattern Analysis Agent"
    model_datetime = _model_datetime(projection)
    workbook.properties.created = model_datetime
    workbook.properties.modified = model_datetime

    section_by_id = {
        str(section.get("id")): section
        for section in (projection.get("sections") or [])
        if isinstance(section, Mapping)
    }
    for section_id in projection.get("section_order") or []:
        section = section_by_id.get(str(section_id))
        if section is None:
            continue
        sheet = workbook.create_sheet(
            SHEET_NAMES.get(str(section_id), str(section_id)[:31])
        )
        _append_section(sheet, section)
        _autosize(sheet)

    output = BytesIO()
    workbook.save(output)
    return _normalize_archive(output.getvalue())
