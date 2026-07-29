"""Evaluation report exporters (PDF / Excel / CSV / JSON)."""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


class EvaluationReportGenerator:
    def __init__(self, *, output_dir: Path | str, formats: list[str] | None = None) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.formats = [f.lower() for f in (formats or ["json", "csv", "excel", "pdf"])]

    def generate(self, report: dict[str, Any]) -> dict[str, str | None]:
        execution_id = report.get("execution_id", "unknown")
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        base = self.output_dir / f"evaluation_{execution_id}_{stamp}"
        paths: dict[str, str | None] = {
            "json": None,
            "csv": None,
            "excel": None,
            "pdf": None,
        }

        if "json" in self.formats:
            json_path = Path(str(base) + ".json")
            json_path.write_text(
                json.dumps(_json_safe(report), indent=2, default=str),
                encoding="utf-8",
            )
            paths["json"] = str(json_path)

        if "csv" in self.formats:
            csv_path = Path(str(base) + "_validation.csv")
            _write_validation_csv(csv_path, report)
            paths["csv"] = str(csv_path)

        if "excel" in self.formats:
            try:
                excel_path = Path(str(base) + ".xlsx")
                _write_excel(excel_path, report)
                paths["excel"] = str(excel_path)
            except Exception:
                paths["excel"] = None

        if "pdf" in self.formats:
            try:
                pdf_path = Path(str(base) + ".pdf")
                _write_pdf(pdf_path, report)
                paths["pdf"] = str(pdf_path)
            except Exception:
                paths["pdf"] = None

        return paths


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _write_validation_csv(path: Path, report: dict[str, Any]) -> None:
    rows: list[dict[str, Any]] = []
    for dataset in report.get("dataset_results", []):
        dataset_id = dataset.get("dataset", {}).get("dataset_id", "")
        for item in dataset.get("validation", []):
            rows.append(
                {
                    "dataset_id": dataset_id,
                    "module": item.get("module"),
                    "status": item.get("status"),
                    "explanation": item.get("explanation"),
                    "duration_ms": item.get("duration_ms"),
                }
            )
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["dataset_id", "module", "status", "explanation", "duration_ms"],
        )
        writer.writeheader()
        writer.writerows(rows)


def _write_excel(path: Path, report: dict[str, Any]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Executive"
    summary.append(["Execution ID", report.get("execution_id")])
    summary.append(["Datasets Evaluated", report.get("datasets_evaluated")])
    for key, value in (report.get("pass_fail_summary") or {}).items():
        summary.append([key, value])

    validation = wb.create_sheet("Validation")
    validation.append(["Dataset", "Module", "Status", "Explanation", "Duration ms"])
    for dataset in report.get("dataset_results", []):
        dataset_id = dataset.get("dataset", {}).get("dataset_id", "")
        for item in dataset.get("validation", []):
            validation.append(
                [
                    dataset_id,
                    item.get("module"),
                    item.get("status"),
                    item.get("explanation"),
                    item.get("duration_ms"),
                ]
            )

    bench = wb.create_sheet("Benchmark")
    bench.append(["Dataset", "Stage", "Avg ms", "Max ms", "Meets Target"])
    for dataset in report.get("dataset_results", []):
        dataset_id = dataset.get("dataset", {}).get("dataset_id", "")
        for stage in dataset.get("benchmark", {}).get("stages", []):
            bench.append(
                [
                    dataset_id,
                    stage.get("name"),
                    stage.get("avg_ms"),
                    stage.get("max_ms"),
                    stage.get("meets_target"),
                ]
            )

    training = wb.create_sheet("Training")
    train = report.get("latest_training") or {}
    for key, value in train.items():
        if isinstance(value, (dict, list)):
            training.append([key, json.dumps(value, default=str)])
        else:
            training.append([key, value])

    wb.save(path)


def _write_pdf(path: Path, report: dict[str, Any]) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "AI Evaluation & Validation Report")
    y -= 24
    c.setFont("Helvetica", 10)
    lines = [
        f"Execution ID: {report.get('execution_id')}",
        f"Datasets evaluated: {report.get('datasets_evaluated')}",
        f"PASS/FAIL summary: {report.get('pass_fail_summary')}",
        f"Generated at: {datetime.now(timezone.utc).isoformat()}",
    ]
    for dataset in report.get("dataset_results", [])[:3]:
        dataset_id = dataset.get("dataset", {}).get("dataset_id")
        lines.append(f"Dataset: {dataset_id}")
        ai = dataset.get("ai_evaluation", {})
        lines.append(
            f"  Accuracy={ai.get('accuracy')} F1={ai.get('f1_score')} "
            f"EngineeringScore={ai.get('engineering_score')}"
        )
        for item in dataset.get("validation", [])[:12]:
            lines.append(
                f"  {item.get('module')}: {item.get('status')} — {item.get('explanation')}"
            )

    for line in lines:
        if y < 60:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 10)
        c.drawString(40, y, str(line)[:110])
        y -= 14
    c.save()
