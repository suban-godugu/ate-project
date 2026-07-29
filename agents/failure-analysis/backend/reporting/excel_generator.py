"""Excel report generation via OpenPyXL."""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any


def export_excel_report(
    *,
    report_id: str,
    summaries: dict[str, Any],
    dashboard: dict[str, Any],
    output_dir: Path,
) -> tuple[Path, float]:
    """Generate multi-sheet Excel engineering report."""
    start = time.perf_counter()
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{report_id}.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export") from exc

    wb = Workbook()
    _write_executive_sheet(wb.active, summaries)
    _add_sheet(wb, "Engineering", _engineering_rows(summaries))
    _add_sheet(wb, "Yield", _yield_rows(summaries))
    _add_sheet(wb, "Lots", _table_rows(summaries.get("lot_summary", [])))
    _add_sheet(wb, "Wafers", _table_rows(summaries.get("wafer_summary", [])))
    _add_sheet(wb, "Failure Modes", _table_rows(summaries.get("top_failure_modes", [])))
    _add_sheet(wb, "Root Cause", _root_cause_rows(summaries))
    _add_sheet(wb, "Actions", _action_rows(summaries))
    _add_sheet(wb, "Dashboard", _dashboard_rows(dashboard))

    wb.save(path)
    elapsed_ms = round((time.perf_counter() - start) * 1000, 2)
    return path, elapsed_ms


def _write_executive_sheet(ws: Any, summaries: dict[str, Any]) -> None:
    from openpyxl.styles import Font

    ws.title = "Executive"
    meta = summaries.get("metadata", {})
    exec_sum = summaries.get("executive_summary", {})

    ws["A1"] = meta.get("report_title", "Failure Analysis Report")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Generated At"
    ws["B3"] = meta.get("generated_at", "")
    ws["A4"] = "Upload ID"
    ws["B4"] = meta.get("upload_id", "")
    ws["A5"] = "Filename"
    ws["B5"] = meta.get("original_filename", "")

    row = 7
    for key, value in exec_sum.items():
        ws.cell(row=row, column=1, value=str(key))
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    ws["A20"] = "Headline"
    ws["A20"].font = Font(bold=True)
    ws["B20"] = exec_sum.get("headline", "")


def _add_sheet(wb: Any, title: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _engineering_rows(summaries: dict[str, Any]) -> list[list[Any]]:
    eng = summaries.get("engineering_summary", {})
    rows = [["Metric", "Value"]]
    for key, value in eng.items():
        if key == "technical_highlights":
            continue
        rows.append([key, str(value)])
    rows.append([])
    rows.append(["Observation", ""])
    for obs in summaries.get("engineering_observations", []):
        rows.append([obs, ""])
    return rows


def _yield_rows(summaries: dict[str, Any]) -> list[list[Any]]:
    yld = summaries.get("yield_summary", {})
    rows = [["Metric", "Value"]]
    for key in ("overall_die_failure_rate", "overall_yield_pct"):
        if key in yld:
            rows.append([key, yld[key]])
    rows.append([])
    rows.append(["Lot", "Failure Rate"])
    for lot in summaries.get("lot_summary", [])[:30]:
        rows.append([lot.get("lot_id"), lot.get("failure_rate")])
    return rows


def _table_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    if not items:
        return [["No data"]]
    headers = list(items[0].keys())
    rows = [headers]
    for item in items:
        rows.append([item.get(h) for h in headers])
    return rows


def _root_cause_rows(summaries: dict[str, Any]) -> list[list[Any]]:
    rc = summaries.get("root_cause_summary", {})
    rows = [
        ["Metric", "Value"],
        ["total_predictions", rc.get("total_predictions")],
        ["average_confidence", rc.get("average_confidence")],
        [],
        ["Scan Chain", "Fault Type", "Root Cause", "Confidence"],
    ]
    for pred in rc.get("predictions", []):
        rows.append(
            [
                pred.get("scan_chain_id"),
                pred.get("predicted_fault_type"),
                pred.get("predicted_root_cause"),
                pred.get("confidence_score"),
            ]
        )
    return rows


def _action_rows(summaries: dict[str, Any]) -> list[list[Any]]:
    rows = [["Priority", "Action", "Source", "Rationale"]]
    for rec in summaries.get("recommended_corrective_actions", []):
        rows.append(
            [
                rec.get("priority", ""),
                rec.get("action", ""),
                rec.get("source", ""),
                rec.get("rationale", ""),
            ]
        )
    return rows


def _dashboard_rows(dashboard: dict[str, Any]) -> list[list[Any]]:
    rows = [["Card", "Value"]]
    for card in dashboard.get("summary_cards", []):
        rows.append([card.get("label"), card.get("value")])
    return rows
